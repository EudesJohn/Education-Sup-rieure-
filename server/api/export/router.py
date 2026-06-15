"""Routeur d'export des résultats — CSV, Excel, PDF."""

import csv
import io
import logging

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse

from core.dependencies import get_current_teacher
from core.db import get_session_by_id, get_session_exams, get_submission_by_exam, get_correction_by_submission

logger = logging.getLogger(__name__)

router = APIRouter()


def _get_session_results(session_id: int, teacher: dict) -> list[dict]:
    """Récupère les résultats d'une session avec vérification du propriétaire."""
    session = get_session_by_id(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session introuvable")

    if session["teacher_id"] != teacher["id"] and teacher["role"] != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")

    exams = get_session_exams(session_id)
    rows = []
    for exam in exams:
        sub = get_submission_by_exam(exam["id"])
        if not sub:
            rows.append({
                "student_name": "N/A",
                "student_number": "N/A",
                "status": "non_soumis",
                "score": "",
                "max_score": "",
                "correction_status": "",
            })
            continue

        corr = get_correction_by_submission(sub["id"])
        rows.append({
            "student_name": sub["student_name"],
            "student_number": sub["student_number"],
            "class_name": sub.get("class_name", ""),
            "university": sub.get("university", ""),
            "status": "soumis",
            "submitted_at": sub.get("submitted_at", ""),
            "score": corr.get("final_score") if corr and corr.get("final_score") is not None else "",
            "max_score": corr["max_score"] if corr else "",
            "correction_status": corr["correction_status"] if corr else "en_attente",
            "ai_score": corr.get("ai_score") if corr and corr.get("ai_score") is not None else "",
            "teacher_score": corr.get("teacher_score") if corr and corr.get("teacher_score") is not None else "",
        })
    return rows


@router.get("/sessions/{session_id}/csv")
def export_csv(
    session_id: int,
    teacher: dict = Depends(get_current_teacher),
):
    """Export des résultats d'une session au format CSV."""
    rows = _get_session_results(session_id, teacher)
    if not rows:
        raise HTTPException(status_code=404, detail="Aucun résultat à exporter")

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f"attachment; filename=session_{session_id}_resultats.csv"},
    )


@router.get("/sessions/{session_id}/excel")
def export_excel(
    session_id: int,
    teacher: dict = Depends(get_current_teacher),
):
    """Export des résultats d'une session au format Excel (.xlsx)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl n'est pas installé")

    rows = _get_session_results(session_id, teacher)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Résultats"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    if rows:
        headers = list(rows[0].keys())
        ws.append([h.replace("_", " ").title() for h in headers])
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row in rows:
            ws.append([row[h] for h in headers])

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 3, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=session_{session_id}_resultats.xlsx"},
    )


@router.get("/sessions/{session_id}/pdf")
def export_pdf(
    session_id: int,
    teacher: dict = Depends(get_current_teacher),
):
    """Export des résultats d'une session au format PDF."""
    try:
        from fpdf import FPDF
    except ImportError:
        raise HTTPException(status_code=500, detail="fpdf2 n'est pas installé. Exécutez : pip install fpdf2")

    rows = _get_session_results(session_id, teacher)
    session = get_session_by_id(session_id)

    class ResultsPDF(FPDF):
        def header(self):
            self.set_font("Helvetica", "B", 16)
            title = session["title"] if session else "Session"
            self.cell(0, 10, f"PEAN - Resultats: {title}", align="C", new_x="LMARGIN", new_y="NEXT")
            self.set_font("Helvetica", "", 10)
            subject = session["subject"] if session else "-"
            grading = session["grading_system"] if session else "20"
            self.cell(0, 6, f"Matiere: {subject} | Systeme: /{grading}", align="C", new_x="LMARGIN", new_y="NEXT")
            self.ln(5)
            self.set_draw_color(37, 99, 235)
            self.set_line_width(0.5)
            self.line(10, self.get_y(), 200, self.get_y())
            self.ln(5)

        def footer(self):
            self.set_y(-15)
            self.set_font("Helvetica", "I", 8)
            self.set_text_color(128)
            self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")

    pdf = ResultsPDF()
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    headers = ["Etudiant", "N°", "Statut", "Note IA", "Note Ens.", "Note Finale"]
    col_widths = [40, 30, 30, 25, 25, 30]

    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(37, 99, 235)
    pdf.set_text_color(255, 255, 255)
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 8, header, border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(0, 0, 0)

    fill = False
    for row in rows:
        if pdf.get_y() > 265:
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_fill_color(37, 99, 235)
            pdf.set_text_color(255, 255, 255)
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 8, header, border=1, fill=True, align="C")
            pdf.ln()
            pdf.set_font("Helvetica", "", 8)
            pdf.set_text_color(0, 0, 0)

        if fill:
            pdf.set_fill_color(249, 250, 251)
        else:
            pdf.set_fill_color(255, 255, 255)

        status_map = {
            "teacher_reviewed": "Validee", "ai_corrected": "Corrigee IA",
            "pending": "En attente", "en_attente": "En attente",
        }
        data = [
            row.get("student_name", "")[:25],
            row.get("student_number", "")[:15],
            status_map.get(row.get("correction_status", ""), row.get("status", "")),
            str(row.get("ai_score", "") or ""),
            str(row.get("teacher_score", "") or ""),
            str(row.get("score", "") or row.get("final_score", "") or ""),
        ]
        for i, val in enumerate(data):
            pdf.cell(col_widths[i], 7, val, border=1, fill=fill, align="C")
        pdf.ln()
        fill = not fill

    pdf.ln(10)
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(37, 99, 235)
    pdf.cell(0, 8, "Resume", new_x="LMARGIN", new_y="NEXT")

    scores = [r.get("score") or r.get("final_score") for r in rows if (r.get("score") or r.get("final_score")) not in (None, "", "N/A")]
    scores_num = [float(s) for s in scores if s]

    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(0, 0, 0)
    total_etudiants = len(rows)
    soumis = sum(1 for r in rows if r.get("status") == "soumis" or r.get("submitted_at"))
    corriges = sum(1 for r in rows if r.get("correction_status") in ("teacher_reviewed", "ai_corrected"))

    pdf.cell(0, 6, f"Total etudiants: {total_etudiants}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"Copies soumises: {soumis}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"Copies corrigees: {corriges}", new_x="LMARGIN", new_y="NEXT")
    if scores_num:
        moyenne = sum(scores_num) / len(scores_num)
        grading = session["grading_system"] if session else "20"
        pdf.cell(0, 6, f"Moyenne: {moyenne:.2f}/{grading}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(0, 6, f"Minimum: {min(scores_num):.2f} | Maximum: {max(scores_num):.2f}", new_x="LMARGIN", new_y="NEXT")

    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=session_{session_id}_resultats.pdf"},
    )
